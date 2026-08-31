
from io import BytesIO
from decimal import Decimal
from uuid import UUID
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.product import Product
from app.services.embedding_service import get_model
from app.schema.product_schema import ProductCreatePayload
from app.services.product_base_service import get_product_base


REQUIRED_COLUMNS = [
    "product_id",
    "name",
    "description",
]

ALL_COLUMNS = [
    "product_id",
    "sku",
    "name",
    "description",
    "category",
    "brand",
    "price",
    "currency",
    "availability",
    "tags",
    "image_url",
]


def normalize_value(value):
    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()

        return value if value else None

    return value


def build_searchable_text(product: dict) -> str:
    return "\n".join(
        [
            f"Product: {product.get('name') or ''}",
            f"Description: {product.get('description') or ''}",
            f"Category: {product.get('category') or ''}",
            f"Brand: {product.get('brand') or ''}",
            f"Keywords: {product.get('tags') or ''}",
        ]
    )

def build_searchable_text_from_playload(product: ProductCreatePayload) -> str:
    return "\n".join(
        [
            f"Product: {product.name or ''}",
            f"Description: {product.description or ''}",
            f"Category: {product.category or ''}",
            f"Brand: {product.brand or ''}",
            f"Keywords: {product.tags or ''}",
        ]
    )

def read_excel(file_bytes: bytes) -> list[dict]:
    workbook = load_workbook(
        filename=BytesIO(file_bytes),
        read_only=True,
        data_only=True,
    )

    worksheet = workbook.active

    rows = worksheet.iter_rows(values_only=True)

    header_row = next(rows, None)

    if not header_row:
        raise ValueError("Excel file is empty.")

    headers = [
        str(value).strip().lower()
        if value is not None
        else ""
        for value in header_row
    ]

    missing_columns = [
        column
        for column in REQUIRED_COLUMNS
        if column not in headers
    ]

    if missing_columns:
        raise ValueError(
            f"Missing required columns: "
            f"{', '.join(missing_columns)}"
        )

    products = []

    for row_number, row in enumerate(rows, start=2):

        if not any(value is not None for value in row):
            continue

        product = {}

        for index, header in enumerate(headers):
            if header in ALL_COLUMNS:
                value = row[index] if index < len(row) else None
                product[header] = normalize_value(value)

        product["_row_number"] = row_number

        products.append(product)

    return products


def validate_products(products: list[dict]) -> list[dict]:
    errors = []

    seen_product_ids = set()

    for product in products:

        row_number = product["_row_number"]

        product_id = product.get("product_id")
        name = product.get("name")
        description = product.get("description")

        if not product_id:
            errors.append(
                {
                    "row": row_number,
                    "field": "product_id",
                    "message": "Product ID is required.",
                }
            )

        elif product_id in seen_product_ids:
            errors.append(
                {
                    "row": row_number,
                    "field": "product_id",
                    "message": f"Duplicate product ID: {product_id}",
                }
            )

        else:
            seen_product_ids.add(product_id)

        if not name:
            errors.append(
                {
                    "row": row_number,
                    "field": "name",
                    "message": "Product name is required.",
                }
            )

        if not description:
            errors.append(
                {
                    "row": row_number,
                    "field": "description",
                    "message": "Product description is required.",
                }
            )

        price = product.get("price")

        if price is not None:
            try:
                Decimal(str(price))
            except Exception:
                errors.append(
                    {
                        "row": row_number,
                        "field": "price",
                        "message": "Price must be a valid number.",
                    }
                )

    return errors

def create_product(
    db: Session,
    tenant_id: UUID,
    product: ProductCreatePayload,
):
    try:
        product_base = get_product_base(
            db=db,
            tenant_id=tenant_id,
        )

        model = get_model()

        searchable_text = build_searchable_text_from_playload(product)

        embedding = model.encode(
            searchable_text,
            show_progress_bar=False,
        ).tolist()

        print(product)

        db_product = Product(
            tenant_id=tenant_id,
            product_base_id=product_base.id,
            product_id=product.product_id,
            sku=product.sku,
            name=product.name,
            description=product.description,
            category=product.category,
            brand=product.brand,
            price=product.price,
            currency=product.currency,
            availability=product.availability,
            tags=product.tags,
            image_url=product.image_url,
            searchable_text=searchable_text,
            embedding=embedding,
        )

        db.add(db_product)
        db.commit()
        db.refresh(db_product)

        return db_product

    except Exception:
        db.rollback()
        raise   

def import_products(
    db: Session,
    file_bytes: bytes,
    tenant_id,
    product_base_id,
):
    products = read_excel(file_bytes)

    errors = validate_products(products)

    if errors:
        return {
            "success": False,
            "errors": errors,
            "imported": 0,
        }

    model = get_model()

    searchable_texts = [
        build_searchable_text(product)
        for product in products
    ]

    embeddings = model.encode(
        searchable_texts,
        batch_size=32,
        show_progress_bar=False,
    )

    db_products = []

    for product, embedding, searchable_text in zip(
        products,
        embeddings,
        searchable_texts,
    ):
        db_product = Product(
            tenant_id=tenant_id,
            product_base_id=product_base_id,
            product_id=product["product_id"],
            sku=product.get("sku"),
            name=product["name"],
            description=product.get("description"),
            category=product.get("category"),
            brand=product.get("brand"),
            price=product.get("price"),
            currency=product.get("currency"),
            availability=product.get("availability"),
            tags=product.get("tags"),
            image_url=product.get("image_url"),
            searchable_text=searchable_text,
            embedding=embedding.tolist(),
        )

        db_products.append(db_product)

    try:
        db.add_all(db_products)
        db.commit()

    except Exception:
        db.rollback()
        raise

    return {
        "success": True,
        "imported": len(db_products),
        "errors": [],
    }


def process_import_job(
    job_id,
    file_bytes,
    tenant_id,
    product_base_id,
):
    from app.core.database import SessionLocal
    from app.models.import_job import ImportJob

    db = SessionLocal()

    try:
        job = db.query(ImportJob).filter(
            ImportJob.id == job_id
        ).first()

        if not job:
            return

        job.status = "processing"
        db.commit()

        # Read Excel
        products = read_excel(file_bytes)

        job.total_rows = len(products)
        db.commit()

        # Validate
        errors = validate_products(products)

        if errors:
            job.status = "failed"
            job.failed_rows = len(errors)
            job.error = str(errors)
            db.commit()
            return

        # Build searchable text
        searchable_texts = [
            build_searchable_text(product)
            for product in products
        ]

        # Generate embeddings in batches
        model = get_model()

        embeddings = model.encode(
            searchable_texts,
            batch_size=32,
            show_progress_bar=False,
        )

        db_products = []

        for index, (
            product,
            embedding,
            searchable_text,
        ) in enumerate(
            zip(
                products,
                embeddings,
                searchable_texts,
            )
        ):
            db_product = Product(
                
                
                tenant_id=tenant_id,
                product_base_id=product_base_id,
                product_id=product["product_id"],
                sku=product.get("sku"),
                name=product["name"],
                description=product.get("description"),
                category=product.get("category"),
                brand=product.get("brand"),
                price=product.get("price"),
                currency=product.get("currency"),
                availability=product.get("availability"),
                tags=product.get("tags"),
                image_url=product.get("image_url"),
                searchable_text=searchable_text,
                embedding=embedding.tolist(),
            )

            db_products.append(db_product)

            # Update progress every 10 products
            if (index + 1) % 10 == 0:
                job.processed_rows = index + 1
                db.commit()

        db.add_all(db_products)
        db.commit()

        job.processed_rows = len(products)
        job.status = "completed"

        from datetime import datetime, timezone

        job.completed_at = datetime.now(timezone.utc)

        db.commit()

    except Exception as error:
        db.rollback()

        job = db.query(ImportJob).filter(
            ImportJob.id == job_id
        ).first()

        if job:
            job.status = "failed"
            job.error = str(error)
            db.commit()

    finally:
        db.close()